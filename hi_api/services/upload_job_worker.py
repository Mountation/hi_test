from db.sqlalchemy import SessionLocal, engine
from db.models import Job as JobORM, EvalData as EvalDataORM
from services.eval_set_service import eval_set_service
import openpyxl
from sqlalchemy import insert
from models.eval_data import EvalDataCreate
from utils.log import get_logger
from sqlalchemy.exc import SQLAlchemyError
from datetime import datetime

logger = get_logger("upload_job_worker")

BATCH_SIZE = 500


def process_upload_job(job_id: str):
    logger.info(f"process_upload_job started for job={job_id}")
    with SessionLocal() as session:
        job = session.query(JobORM).filter(JobORM.job_id == job_id).first()
        if not job:
            logger.error(f"job {job_id} not found")
            return
        try:
            logger.info(f"DB engine url: {engine.url}")
        except Exception:
            logger.info("DB engine url: <unavailable>")
        logger.info(f"upload job details: eval_set_id={job.eval_set_id}, file_path={job.file_path}")
        # mark running
        job.status = 'running'
        job.started_at = datetime.utcnow()
        session.add(job)
        session.commit()
        file_path = job.file_path

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        # skip header
        try:
            header = next(rows_iter)
        except StopIteration:
            header = None
        batch = []
        total = 0
        processed = 0
        skipped = 0
        # compute starting corpus_id for this eval_set (max existing corpus_id)
        try:
            with SessionLocal() as s2:
                max_row = s2.query(EvalDataORM).filter(EvalDataORM.eval_set_id == job.eval_set_id).order_by(EvalDataORM.corpus_id.desc()).with_entities(EvalDataORM.corpus_id).first()
                start_corpus = int(max_row[0]) if (max_row and max_row[0] is not None) else 0
        except Exception:
            start_corpus = 0
        # Two-pass approach: first count non-empty rows to get an accurate total, then re-iterate to insert.
        # First pass: count non-empty rows (skip header)
        total_expected = 0
        try:
            count_iter = ws.iter_rows(values_only=True)
            _ = next(count_iter, None)
            for crow in count_iter:
                ccontent = crow[0] if len(crow) > 0 else None
                if ccontent is None or (isinstance(ccontent, str) and ccontent.strip() == ''):
                    continue
                total_expected += 1
        except Exception as e:
            logger.warning(f"failed to compute total_expected by scanning rows: {e}")

        # persist total_expected to job record so frontend sees a stable total
        try:
            with SessionLocal() as s_tot:
                j = s_tot.query(JobORM).filter(JobORM.job_id == job_id).first()
                if j:
                    j.total = total_expected
                    s_tot.add(j)
                    s_tot.commit()
        except Exception as e:
            logger.warning(f"failed to persist total_expected for job {job_id}: {e}")

        # Recreate iterator for insertion pass and skip header
        rows_iter = ws.iter_rows(values_only=True)
        try:
            _ = next(rows_iter)
        except StopIteration:
            rows_iter = iter(())

        # Use total_expected as the stable total for progress updates
        total = total_expected
        for row in rows_iter:
            content = row[0] if len(row) > 0 else None
            # treat numeric 0 as valid content; skip only None or whitespace-only strings
            if content is None or (isinstance(content, str) and content.strip() == ''):
                skipped += 1
                logger.debug(f"skipping empty row (content={repr(content)}), skipped_count={skipped}")
                continue
            # we already counted non-empty rows in total_expected
            expected = row[1] if len(row) > 1 else None
            intent = row[2] if len(row) > 2 else None
            # assign corpus_id sequentially
            start_corpus += 1
            batch.append({
                'eval_set_id': job.eval_set_id,
                'corpus_id': start_corpus,
                'content': str(content),
                'expected': str(expected) if expected is not None else None,
                'intent': str(intent) if intent is not None else None,
                'deleted': False
            })
            if len(batch) >= BATCH_SIZE:
                logger.info(f"inserting batch of size={len(batch)} for job={job_id}")
                _bulk_insert_batch(batch)
                processed += len(batch)
                batch = []
                _update_job_progress(job_id, processed, total_expected)
                # quick verification: count rows in DB for this eval_set_id
                try:
                    with SessionLocal() as scheck:
                        cnt = scheck.query(EvalDataORM).filter(EvalDataORM.eval_set_id == job.eval_set_id, EvalDataORM.deleted == False).count()
                        logger.info(f"post-insert check: eval_data rows for set {job.eval_set_id} = {cnt}")
                except Exception as e:
                    logger.warning(f"post-insert count check failed: {e}")
        # final batch
        if batch:
            logger.info(f"inserting final batch of size={len(batch)} for job={job_id}")
            _bulk_insert_batch(batch)
            processed += len(batch)
            _update_job_progress(job_id, processed, total_expected)
            try:
                with SessionLocal() as scheck:
                    cnt = scheck.query(EvalDataORM).filter(EvalDataORM.eval_set_id == job.eval_set_id, EvalDataORM.deleted == False).count()
                    logger.info(f"post-final-insert check: eval_data rows for set {job.eval_set_id} = {cnt}")
            except Exception as e:
                logger.warning(f"post-final-insert count check failed: {e}")

        logger.info(f"upload summary for job={job_id}: processed={processed}, total={total}, skipped={skipped}")

        # final refresh count
        try:
            eval_set_service.refresh_count(job.eval_set_id)
        except Exception as e:
            logger.warning(f"refresh_count failed after job {job_id}: {e}")

        with SessionLocal() as session:
            job = session.query(JobORM).filter(JobORM.job_id == job_id).first()
            job.status = 'success'
            job.processed = processed
            job.total = total
            job.finished_at = datetime.utcnow()
            session.add(job)
            session.commit()
        logger.info(f"process_upload_job finished for job={job_id}, processed={processed}")
    except Exception as e:
        logger.exception(f"process_upload_job failed for job={job_id}: {e}")
        with SessionLocal() as session:
            job = session.query(JobORM).filter(JobORM.job_id == job_id).first()
            if job:
                job.status = 'failed'
                job.error = str(e)
                job.finished_at = datetime.utcnow()
                session.add(job)
                session.commit()


def _bulk_insert_batch(batch):
    # Use SQLAlchemy core bulk insert for speed
    with SessionLocal() as session:
        try:
            # use the ORM table metadata explicitly to ensure insert targets the correct table
            session.execute(insert(EvalDataORM.__table__), batch)
            session.commit()
        except SQLAlchemyError as e:
            session.rollback()
            # Bulk insert failed — log and attempt a safer per-row fallback to avoid losing data
            logger.exception(f"bulk insert failed: {e}; batch_size={len(batch)}. Attempting per-row fallback.")
            try:
                for row in batch:
                    obj = EvalDataORM(**row)
                    session.add(obj)
                session.commit()
                logger.info(f"per-row fallback succeeded, inserted {len(batch)} rows")
            except Exception as e2:
                session.rollback()
                logger.exception(f"per-row fallback insert failed: {e2}; batch_size={len(batch)}")


def _update_job_progress(job_id: str, processed: int, total: int):
    with SessionLocal() as session:
        job = session.query(JobORM).filter(JobORM.job_id == job_id).first()
        if not job:
            return
        job.processed = processed
        job.total = total
        session.add(job)
        session.commit()
